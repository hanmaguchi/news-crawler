from datetime import datetime
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from crawler.export import to_csv, to_excel, to_json
from crawler.models import Article

_ART = Article(
    title="테스트 기사",
    press="연합뉴스",
    pub_date=datetime(2024, 5, 1, 9, 30),
    url="https://example.com/news/1",
    source="네이버",
)

_ART_POS = Article(
    title="삼성전자 주가 급등 신고가 달성",
    press="한국경제",
    pub_date=datetime(2024, 5, 1),
    url="https://example.com/2",
    source="구글",
)


def test_link_cell_has_hyperlink_and_tooltip():
    # 컬럼 순서: 순번(1) 배포일자(2) 언론사(3) 제목(4) 링크(5) 감성(6) 출처(7)
    with patch("crawler.export.classify", return_value="중립"):
        wb = load_workbook(BytesIO(to_excel([_ART])))
    ws = wb.active
    link_cell = ws.cell(row=2, column=5)
    assert link_cell.value == "기사보기"
    assert link_cell.hyperlink.target == "https://example.com/news/1"
    assert link_cell.hyperlink.tooltip == "https://example.com/news/1"


def test_headers():
    with patch("crawler.export.classify", return_value="중립"):
        wb = load_workbook(BytesIO(to_excel([])))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["순번", "배포일자", "언론사", "제목", "링크", "감성", "출처"]


def test_sentiment_column_filled():
    with patch("crawler.export.classify", return_value="긍정"):
        wb = load_workbook(BytesIO(to_excel([_ART_POS])))
    ws = wb.active
    assert ws.cell(row=2, column=6).value == "긍정"


def test_to_csv_returns_utf8_bom():
    with patch("crawler.export.classify", return_value="중립"):
        data = to_csv([_ART])
    assert data[:3] == b"\xef\xbb\xbf"  # UTF-8 BOM
    text = data.decode("utf-8-sig")
    assert "테스트 기사" in text
    assert "연합뉴스" in text


def test_to_json_structure():
    with patch("crawler.export.classify", return_value="긍정"):
        import json
        data = json.loads(to_json([_ART_POS]).decode("utf-8"))
    assert len(data) == 1
    assert data[0]["순번"] == 1
    assert data[0]["제목"] == _ART_POS.title
    assert data[0]["감성"] == "긍정"
