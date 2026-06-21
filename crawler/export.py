import csv
import json
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from crawler.models import Article
from crawler.sentiment import classify

_HEADERS = ["순번", "배포일자", "언론사", "제목", "링크", "감성", "출처"]
_WIDTHS = [6, 18, 16, 60, 12, 10, 14]
_LINK_FONT = Font(color="0563C1", underline="single")

_FILL = {
    "긍정": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "부정": PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "중립": None,
}


def to_excel(articles: list[Article], sentiments: list[str] | None = None) -> bytes:
    if sentiments is None:
        sentiments = [classify(a.title) for a in articles]

    wb = Workbook()
    ws = wb.active
    ws.title = "뉴스"

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for idx, (art, sentiment) in enumerate(zip(articles, sentiments), start=1):
        ws.append([idx, art.pub_date, art.press, art.title, "기사보기", sentiment, art.source])
        row = ws.max_row

        ws.cell(row=row, column=2).number_format = "yyyy-mm-dd hh:mm"

        sent_cell = ws.cell(row=row, column=6)
        fill = _FILL.get(sentiment)
        if fill:
            sent_cell.fill = fill

        link_cell = ws.cell(row=row, column=5)
        if art.url:
            link_cell.hyperlink = Hyperlink(
                ref=link_cell.coordinate, target=art.url, tooltip=art.url
            )
            link_cell.font = _LINK_FONT

    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_CSV_HEADERS = ["순번", "배포일자", "언론사", "제목", "URL", "감성", "출처"]


def to_csv(articles: list[Article], sentiments: list[str] | None = None) -> bytes:
    """UTF-8 BOM CSV — Excel에서 한글 깨짐 없이 열림."""
    if sentiments is None:
        sentiments = [classify(a.title) for a in articles]
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(_CSV_HEADERS)
    for idx, (art, sentiment) in enumerate(zip(articles, sentiments), 1):
        date_str = art.pub_date.strftime("%Y-%m-%d %H:%M") if art.pub_date else ""
        writer.writerow([
            idx, date_str, art.press, art.title,
            art.url, sentiment, art.source,
        ])
    return buf.getvalue().encode("utf-8-sig")


def to_json(articles: list[Article], sentiments: list[str] | None = None) -> bytes:
    if sentiments is None:
        sentiments = [classify(a.title) for a in articles]
    data = [
        {
            "순번": idx,
            "배포일자": art.pub_date.isoformat() if art.pub_date else None,
            "언론사": art.press,
            "제목": art.title,
            "url": art.url,
            "감성": sentiment,
            "출처": art.source,
        }
        for idx, (art, sentiment) in enumerate(zip(articles, sentiments), 1)
    ]
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
